#!/usr/bin/env python
# Written by Yuen-Hsien Tseng on 2018/04/23, modified on 2019/11/18
# run:
# $ python Excel2txt.py Chinese_Humor_Multi-Labeled.xlsx 1 > mlabel_corpora/JokeHumorLevel.txt
# $ python Excel2txt.py Chinese_Humor_Multi-Labeled.xlsx 2 > mlabel_corpora/JokeSkill.txt
# $ python Excel2txt.py Chinese_Humor_Multi-Labeled.xlsx 3 > mlabel_corpora/JokeMotivation.txt
#       3365 records
import sys, re
from openpyxl import load_workbook
f, flag = sys.argv[1:3]

if flag == '1':
    print("ID\tTitle\tContent\tHumorLevel")
elif flag == '2':
    print("ID\tTitle\tContent\t"
    + "Pun\tExaggeration\tAnthropomorphism\t"
    + "Bridge_Inference\tIllogical\tIrony\tImitation\tOthers")
elif flag == '3':
    print("ID\tTitle\tContent\t"
    + "Affinity\tSelf_Improvement\tAttack\tSelf_Depression\tTaboo\tOthers")
else:
    print("Syntax: program filename.xlsx [1|2|3]")

n = 0
wb = load_workbook(filename = f)
#print("Number of worksheets: {}".format(len(wb.sheetnames)))
for nwb, ws in enumerate(wb):
#    print(ws.title)
    for i, r in enumerate(ws.rows):
        if i == 0: continue # skip column titles
# ID	標題	內容	好笑程度	雙關	誇飾	擬人化	橋介推論	法則誤用	悖理	模仿	其他	親和	自我提升	攻擊	自我貶抑	禁忌	其他
        ID = r[0].value # get joke ID
        if ID == None: continue # skip empty ID
        if re.match(r'[^L]\d+', ID): continue # skip if not start with L 
        r[2].value = re.sub(r"[\n\r]", "     ", r[2].value)
        if flag == '1':
            T = [r[i].value if i<3 else str(r[i].value) for i in (1, 2, 3)]
        elif flag == '2':
            T = [r[i].value if i<3 else str(r[i].value) for i in (1, 2, 4, 5, 6, 7, 8, 9, 10, 11)]
        elif flag == '3':
            T = [r[i].value if i<3 else str(r[i].value) for i in (1, 2, 12, 13, 14, 15, 16, 17)]
        print(ID + "\t" + "\t".join(T))
        n += 1
        #if i>4: break
#    if nwb>1: break
sys.stderr.write(f"{n} records\n")
